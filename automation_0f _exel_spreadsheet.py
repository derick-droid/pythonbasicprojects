import openpyxl as xl
from openpyxl.chart import BarChart, Reference
# cell = sheet.cell(1, 1)
# cell = sheet["a1"]
# print(cell.value)
# print(sheet.max_row)  # to know the number of rows present in our spread sheet
# for row in range(1, sheet.max_row + 1):  # we add one to capture the fourth row
#     print(row)


def process_working(filename):

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_prices = cell.value * 0.9
        corrected_prices_cell = sheet.cell(row, 4)
        corrected_prices_cell.value = corrected_prices
    value = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(value)
    sheet.add_chart(chart, "e2")

    wb.save(filename)


